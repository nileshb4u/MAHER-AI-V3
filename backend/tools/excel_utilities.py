"""
Excel Utilities Tool for MAHER Orchestrator

Provides comprehensive Excel manipulation capabilities including:
- Read Excel files
- Modify existing workbooks
- Generate new Excel files
- Create pivot tables
- Generate charts and visualizations
- Data analysis and formatting
"""

from typing import Dict, Any, List, Optional, Union
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.utils import get_column_letter
    import pandas as pd
    import xlsxwriter
    DEPENDENCIES_AVAILABLE = True
except ImportError as e:
    DEPENDENCIES_AVAILABLE = False
    IMPORT_ERROR = str(e)


def read_excel(
    input_file: str,
    sheet_name: Optional[str] = None,
    cell_range: Optional[str] = None,
    as_dataframe: bool = False
) -> Dict[str, Any]:
    """
    Read data from Excel file

    Args:
        input_file: Path to Excel file
        sheet_name: Specific sheet to read (None for first sheet)
        cell_range: Specific range to read (e.g., 'A1:D10')
        as_dataframe: Return as pandas DataFrame

    Returns:
        Excel data
    """
    if not DEPENDENCIES_AVAILABLE:
        return {
            "success": False,
            "error": f"Required dependencies not available: {IMPORT_ERROR}"
        }

    try:
        if as_dataframe:
            # Use pandas for DataFrame
            df = pd.read_excel(input_file, sheet_name=sheet_name or 0)
            return {
                "success": True,
                "data": df.to_dict('records'),
                "columns": list(df.columns),
                "rows": len(df),
                "shape": df.shape,
                "message": "Excel data loaded as DataFrame"
            }
        else:
            # Use openpyxl for raw data
            wb = load_workbook(input_file)
            ws = wb[sheet_name] if sheet_name else wb.active

            # Extract data
            if cell_range:
                data = []
                for row in ws[cell_range]:
                    data.append([cell.value for cell in row])
            else:
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(list(row))

            return {
                "success": True,
                "data": data,
                "sheet_name": ws.title,
                "rows": len(data),
                "columns": len(data[0]) if data else 0,
                "available_sheets": wb.sheetnames,
                "message": "Excel data loaded successfully"
            }

    except Exception as e:
        return {
            "success": False,
            "error": f"Failed to read Excel: {str(e)}"
        }


def create_excel(
    output_file: str,
    data: Union[List[List[Any]], Dict[str, List[List[Any]]]],
    headers: Optional[List[str]] = None,
    sheet_name: str = "Sheet1",
    formatting: Optional[Dict[str, Any]] = None
) -> Dict[str, Any]:
    """
    Create a new Excel file

    Args:
        output_file: Path for output Excel file
        data: 2D list of data or dict of {sheet_name: data}
        headers: Column headers
        sheet_name: Name for the sheet (if data is list)
        formatting: Formatting options (font, colors, etc.)

    Returns:
        Result dictionary
    """
    if not DEPENDENCIES_AVAILABLE:
        return {
            "success": False,
            "error": f"Required dependencies not available: {IMPORT_ERROR}"
        }

    try:
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Handle multi-sheet data
        if isinstance(data, dict):
            for sheet_name, sheet_data in data.items():
                ws = wb.create_sheet(sheet_name)
                _write_data_to_sheet(ws, sheet_data, headers, formatting)
        else:
            ws = wb.create_sheet(sheet_name)
            _write_data_to_sheet(ws, data, headers, formatting)

        wb.save(output_file)

        return {
            "success": True,
            "output_file": output_file,
            "sheets": wb.sheetnames,
            "message": "Excel file created successfully"
        }

    except Exception as e:
        return {
            "success": False,
            "error": f"Failed to create Excel: {str(e)}"
        }


def _write_data_to_sheet(
    ws,
    data: List[List[Any]],
    headers: Optional[List[str]] = None,
    formatting: Optional[Dict[str, Any]] = None
):
    """Helper function to write data to a worksheet"""
    row_offset = 1

    # Add headers if provided
    if headers:
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)

            # Apply header formatting
            if formatting and formatting.get("header_style"):
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)

        row_offset = 2

    # Write data
    for row_idx, row_data in enumerate(data, row_offset):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    if formatting and formatting.get("auto_width"):
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)


def modify_excel(
    input_file: str,
    output_file: str,
    modifications: List[Dict[str, Any]]
) -> Dict[str, Any]:
    """
    Modify existing Excel file

    Args:
        input_file: Path to input Excel file
        output_file: Path for output file
        modifications: List of modifications
            Example: [
                {"action": "set_cell", "sheet": "Sheet1", "cell": "A1", "value": "New"},
                {"action": "add_row", "sheet": "Sheet1", "data": [1, 2, 3]},
                {"action": "delete_row", "sheet": "Sheet1", "row": 5}
            ]

    Returns:
        Result dictionary
    """
    if not DEPENDENCIES_AVAILABLE:
        return {
            "success": False,
            "error": f"Required dependencies not available: {IMPORT_ERROR}"
        }

    try:
        wb = load_workbook(input_file)

        for mod in modifications:
            action = mod.get("action", "")
            sheet_name = mod.get("sheet")
            ws = wb[sheet_name] if sheet_name else wb.active

            if action == "set_cell":
                cell = mod.get("cell", "A1")
                value = mod.get("value", "")
                ws[cell] = value

            elif action == "add_row":
                data = mod.get("data", [])
                ws.append(data)

            elif action == "delete_row":
                row_num = mod.get("row", 1)
                ws.delete_rows(row_num)

            elif action == "delete_column":
                col_num = mod.get("column", 1)
                ws.delete_cols(col_num)

            elif action == "insert_row":
                row_num = mod.get("row", 1)
                ws.insert_rows(row_num)

            elif action == "format_cell":
                cell = mod.get("cell", "A1")
                cell_obj = ws[cell]

                if mod.get("bold"):
                    cell_obj.font = Font(bold=True)
                if mod.get("color"):
                    cell_obj.fill = PatternFill(start_color=mod["color"], fill_type="solid")

        wb.save(output_file)

        return {
            "success": True,
            "output_file": output_file,
            "modifications_applied": len(modifications),
            "message": "Excel file modified successfully"
        }

    except Exception as e:
        return {
            "success": False,
            "error": f"Failed to modify Excel: {str(e)}"
        }


def create_pivot_table(
    input_file: str,
    output_file: str,
    sheet_name: str,
    pivot_config: Dict[str, Any]
) -> Dict[str, Any]:
    """
    Create pivot table from Excel data

    Args:
        input_file: Path to input Excel file
        output_file: Path for output file with pivot
        sheet_name: Source data sheet name
        pivot_config: Pivot table configuration
            {
                "rows": ["Category"],
                "columns": ["Month"],
                "values": {"Sales": "sum"},
                "output_sheet": "Pivot"
            }

    Returns:
        Result dictionary
    """
    if not DEPENDENCIES_AVAILABLE:
        return {
            "success": False,
            "error": f"Required dependencies not available: {IMPORT_ERROR}"
        }

    try:
        # Read data with pandas
        df = pd.read_excel(input_file, sheet_name=sheet_name)

        # Create pivot table
        rows = pivot_config.get("rows", [])
        columns = pivot_config.get("columns", [])
        values = pivot_config.get("values", {})

        pivot = pd.pivot_table(
            df,
            index=rows,
            columns=columns,
            values=list(values.keys()),
            aggfunc=list(values.values())[0] if values else 'sum'
        )

        # Save to Excel
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Copy original data
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Add pivot table
            pivot_sheet = pivot_config.get("output_sheet", "Pivot")
            pivot.to_excel(writer, sheet_name=pivot_sheet)

        return {
            "success": True,
            "output_file": output_file,
            "pivot_shape": pivot.shape,
            "message": "Pivot table created successfully"
        }

    except Exception as e:
        return {
            "success": False,
            "error": f"Failed to create pivot table: {str(e)}"
        }


def add_chart_to_excel(
    input_file: str,
    output_file: str,
    chart_config: Dict[str, Any]
) -> Dict[str, Any]:
    """
    Add chart to Excel file

    Args:
        input_file: Path to input Excel file
        output_file: Path for output file
        chart_config: Chart configuration
            {
                "type": "bar",  # bar, line, pie
                "sheet": "Sheet1",
                "data_range": "A1:B10",
                "title": "Sales Chart",
                "position": "D2"
            }

    Returns:
        Result dictionary
    """
    if not DEPENDENCIES_AVAILABLE:
        return {
            "success": False,
            "error": f"Required dependencies not available: {IMPORT_ERROR}"
        }

    try:
        wb = load_workbook(input_file)
        sheet_name = chart_config.get("sheet", wb.sheetnames[0])
        ws = wb[sheet_name]

        chart_type = chart_config.get("type", "bar")
        data_range = chart_config.get("data_range", "A1:B10")
        title = chart_config.get("title", "Chart")
        position = chart_config.get("position", "D2")

        # Parse data range
        data_ref = Reference(ws, range_string=f"{sheet_name}!{data_range}")

        # Create chart based on type
        if chart_type == "bar":
            chart = BarChart()
        elif chart_type == "line":
            chart = LineChart()
        elif chart_type == "pie":
            chart = PieChart()
        else:
            chart = BarChart()

        chart.add_data(data_ref, titles_from_data=True)
        chart.title = title

        ws.add_chart(chart, position)
        wb.save(output_file)

        return {
            "success": True,
            "output_file": output_file,
            "chart_type": chart_type,
            "message": f"{chart_type.title()} chart added successfully"
        }

    except Exception as e:
        return {
            "success": False,
            "error": f"Failed to add chart: {str(e)}"
        }


def analyze_excel_data(
    input_file: str,
    sheet_name: Optional[str] = None,
    analysis_type: str = "summary"
) -> Dict[str, Any]:
    """
    Perform data analysis on Excel file

    Args:
        input_file: Path to Excel file
        sheet_name: Sheet to analyze
        analysis_type: Type of analysis - 'summary', 'statistics', 'describe'

    Returns:
        Analysis results
    """
    if not DEPENDENCIES_AVAILABLE:
        return {
            "success": False,
            "error": f"Required dependencies not available: {IMPORT_ERROR}"
        }

    try:
        df = pd.read_excel(input_file, sheet_name=sheet_name or 0)

        if analysis_type == "summary":
            result = {
                "rows": len(df),
                "columns": len(df.columns),
                "column_names": list(df.columns),
                "data_types": df.dtypes.astype(str).to_dict(),
                "null_counts": df.isnull().sum().to_dict(),
                "memory_usage": df.memory_usage(deep=True).sum()
            }

        elif analysis_type == "statistics":
            result = df.describe().to_dict()

        elif analysis_type == "describe":
            result = {
                "statistics": df.describe().to_dict(),
                "info": {
                    "columns": list(df.columns),
                    "dtypes": df.dtypes.astype(str).to_dict(),
                    "shape": df.shape
                }
            }

        else:
            result = {}

        return {
            "success": True,
            "analysis": result,
            "analysis_type": analysis_type,
            "message": "Analysis completed successfully"
        }

    except Exception as e:
        return {
            "success": False,
            "error": f"Analysis failed: {str(e)}"
        }


def excel_to_csv(
    input_file: str,
    output_file: str,
    sheet_name: Optional[str] = None
) -> Dict[str, Any]:
    """
    Convert Excel to CSV

    Args:
        input_file: Path to Excel file
        output_file: Path for CSV output
        sheet_name: Sheet to convert (None for first sheet)

    Returns:
        Result dictionary
    """
    if not DEPENDENCIES_AVAILABLE:
        return {
            "success": False,
            "error": f"Required dependencies not available: {IMPORT_ERROR}"
        }

    try:
        df = pd.read_excel(input_file, sheet_name=sheet_name or 0)
        df.to_csv(output_file, index=False)

        return {
            "success": True,
            "output_file": output_file,
            "rows": len(df),
            "columns": len(df.columns),
            "message": "Excel converted to CSV successfully"
        }

    except Exception as e:
        return {
            "success": False,
            "error": f"Conversion failed: {str(e)}"
        }


def format_excel(
    input_file: str,
    output_file: str,
    formatting_rules: List[Dict[str, Any]]
) -> Dict[str, Any]:
    """
    Apply formatting to Excel file

    Args:
        input_file: Path to input Excel file
        output_file: Path for output file
        formatting_rules: List of formatting rules
            Example: [
                {"range": "A1:Z1", "bold": true, "bg_color": "366092"},
                {"range": "A2:A100", "number_format": "#,##0.00"}
            ]

    Returns:
        Result dictionary
    """
    if not DEPENDENCIES_AVAILABLE:
        return {
            "success": False,
            "error": f"Required dependencies not available: {IMPORT_ERROR}"
        }

    try:
        wb = load_workbook(input_file)
        ws = wb.active

        for rule in formatting_rules:
            cell_range = rule.get("range", "A1")

            for row in ws[cell_range]:
                for cell in row:
                    if rule.get("bold"):
                        cell.font = Font(bold=True)

                    if rule.get("bg_color"):
                        cell.fill = PatternFill(
                            start_color=rule["bg_color"],
                            end_color=rule["bg_color"],
                            fill_type="solid"
                        )

                    if rule.get("number_format"):
                        cell.number_format = rule["number_format"]

                    if rule.get("alignment"):
                        cell.alignment = Alignment(horizontal=rule["alignment"])

        wb.save(output_file)

        return {
            "success": True,
            "output_file": output_file,
            "rules_applied": len(formatting_rules),
            "message": "Formatting applied successfully"
        }

    except Exception as e:
        return {
            "success": False,
            "error": f"Formatting failed: {str(e)}"
        }


def get_excel_info(
    input_file: str
) -> Dict[str, Any]:
    """
    Get metadata and information about Excel file

    Args:
        input_file: Path to Excel file

    Returns:
        Excel metadata
    """
    if not DEPENDENCIES_AVAILABLE:
        return {
            "success": False,
            "error": f"Required dependencies not available: {IMPORT_ERROR}"
        }

    try:
        wb = load_workbook(input_file)

        sheets_info = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheets_info.append({
                "name": sheet_name,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "dimensions": ws.dimensions
            })

        return {
            "success": True,
            "file_path": input_file,
            "file_size": os.path.getsize(input_file),
            "sheets": sheets_info,
            "total_sheets": len(wb.sheetnames),
            "properties": {
                "creator": wb.properties.creator or "N/A",
                "title": wb.properties.title or "N/A",
                "created": str(wb.properties.created) if wb.properties.created else "N/A",
                "modified": str(wb.properties.modified) if wb.properties.modified else "N/A"
            }
        }

    except Exception as e:
        return {
            "success": False,
            "error": f"Failed to get Excel info: {str(e)}"
        }


def excel_to_pdf(
    input_file: str,
    output_file: str,
    sheet_name: Optional[str] = None,
    include_all_sheets: bool = False
) -> Dict[str, Any]:
    """
    Convert Excel file to PDF format

    Args:
        input_file: Path to Excel file (.xlsx, .xls)
        output_file: Path for output PDF file
        sheet_name: Specific sheet to convert (None for active sheet)
        include_all_sheets: Convert all sheets to multi-page PDF

    Returns:
        Result dictionary with success status and output path
    """
    if not DEPENDENCIES_AVAILABLE:
        return {
            "success": False,
            "error": f"Required dependencies not available: {IMPORT_ERROR}"
        }

    try:
        # Import PDF generation libraries
        try:
            from reportlab.lib.pagesizes import letter, landscape, A4
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, PageBreak, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        except ImportError as e:
            return {
                "success": False,
                "error": f"ReportLab not installed. Install with: pip install reportlab. Error: {str(e)}"
            }

        # Read Excel file
        if include_all_sheets:
            # Load all sheets
            excel_file = pd.ExcelFile(input_file)
            sheet_names = excel_file.sheet_names
            dfs = {sheet: pd.read_excel(excel_file, sheet_name=sheet) for sheet in sheet_names}
        else:
            # Load single sheet
            if sheet_name:
                df = pd.read_excel(input_file, sheet_name=sheet_name)
                dfs = {sheet_name: df}
            else:
                df = pd.read_excel(input_file)
                dfs = {"Sheet": df}

        # Create PDF
        doc = SimpleDocTemplate(
            output_file,
            pagesize=landscape(A4),  # Use landscape for better Excel table fit
            topMargin=0.5*inch,
            bottomMargin=0.5*inch,
            leftMargin=0.5*inch,
            rightMargin=0.5*inch
        )

        elements = []
        styles = getSampleStyleSheet()

        # Style for sheet titles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=12
        )

        for sheet_name_key, df in dfs.items():
            # Add sheet title if multiple sheets
            if include_all_sheets or len(dfs) > 1:
                elements.append(Paragraph(f"Sheet: {sheet_name_key}", title_style))
                elements.append(Spacer(1, 0.2*inch))

            # Convert DataFrame to list of lists for ReportLab Table
            # Include column headers
            data = [list(df.columns)] + df.fillna('').values.tolist()

            # Limit data size for very large Excel files
            MAX_ROWS = 500
            if len(data) > MAX_ROWS:
                data = data[:MAX_ROWS]
                elements.append(Paragraph(
                    f"Note: Showing first {MAX_ROWS-1} rows (file has {len(df)} rows)",
                    styles['Normal']
                ))
                elements.append(Spacer(1, 0.1*inch))

            # Create table
            table = Table(data)

            # Apply table styling
            table.setStyle(TableStyle([
                # Header row
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472c4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),

                # Data rows
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f2f2f2')]),
            ]))

            elements.append(table)

            # Add page break if multiple sheets
            if include_all_sheets and sheet_name_key != list(dfs.keys())[-1]:
                elements.append(PageBreak())

        # Build PDF
        doc.build(elements)

        return {
            "success": True,
            "output_file": output_file,
            "input_file": input_file,
            "sheets_converted": len(dfs),
            "message": f"Excel file converted to PDF successfully ({len(dfs)} sheet(s))"
        }

    except Exception as e:
        return {
            "success": False,
            "error": f"Failed to convert Excel to PDF: {str(e)}"
        }
